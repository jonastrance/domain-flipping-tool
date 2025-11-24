#!/usr/bin/env python3
"""
Domain Pricing Spreadsheet Generator

Creates an Excel spreadsheet template for tracking domain purchases and sales.
Includes formulas for profit calculation and ROI analysis.

Usage:
    python create_pricing_template.py [--output FILENAME]
"""

import argparse
import sys
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: Required packages not installed.")
    print("Please run: pip install -r requirements.txt")
    sys.exit(1)


def create_pricing_template(filename: str = "domain_pricing_template.xlsx"):
    """
    Create an Excel template for domain pricing and tracking.
    
    Args:
        filename: Output filename
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # Create sheets
    inventory_sheet = wb.create_sheet("Domain Inventory", 0)
    sold_sheet = wb.create_sheet("Sold Domains", 1)
    summary_sheet = wb.create_sheet("Summary", 2)
    
    # Configure Domain Inventory sheet
    setup_inventory_sheet(inventory_sheet)
    
    # Configure Sold Domains sheet
    setup_sold_sheet(sold_sheet)
    
    # Configure Summary sheet
    setup_summary_sheet(summary_sheet)
    
    # Save the workbook
    wb.save(filename)
    print(f"✓ Created pricing template: {filename}")
    print(f"\nTemplate includes:")
    print(f"  - Domain Inventory: Track purchased domains and target prices")
    print(f"  - Sold Domains: Record sales and calculate profits")
    print(f"  - Summary: Overview of portfolio performance")


def setup_inventory_sheet(sheet):
    """Set up the Domain Inventory sheet."""
    # Headers
    headers = [
        "Domain Name",
        "Purchase Date",
        "Purchase Price ($)",
        "Registrar",
        "Expiry Date",
        "Target Sale Price ($)",
        "Min Acceptable Price ($)",
        "Listed On",
        "Listing URL",
        "Status",
        "Notes"
    ]
    
    # Style configuration
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    column_widths = [25, 12, 15, 12, 12, 18, 20, 15, 30, 12, 30]
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Add sample data
    sample_data = [
        ["example-domain.com", "2025-01-15", 2.99, "Namecheap", "2026-01-15", 50.00, 25.00, "Flippa", "https://flippa.com/...", "Listed", "Tech related"],
        ["quick-startup.io", "2025-01-16", 8.99, "Namecheap", "2026-01-16", 100.00, 60.00, "Brandpa", "https://brandpa.com/...", "Listed", "Business name"],
        ["smart-hub.net", "2025-01-17", 1.99, "Namecheap", "2026-01-17", 40.00, 20.00, "", "", "Available", "Short and catchy"],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [3, 6, 7]:  # Price columns
                cell.number_format = '$#,##0.00'
    
    # Freeze top row
    sheet.freeze_panes = 'A2'


def setup_sold_sheet(sheet):
    """Set up the Sold Domains sheet."""
    # Headers
    headers = [
        "Domain Name",
        "Purchase Date",
        "Purchase Price ($)",
        "Sale Date",
        "Sale Price ($)",
        "Marketplace",
        "Marketplace Fee ($)",
        "Net Profit ($)",
        "ROI (%)",
        "Days Held",
        "Notes"
    ]
    
    # Style configuration
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    column_widths = [25, 12, 15, 12, 15, 15, 18, 15, 12, 12, 30]
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Add formula rows (starting from row 2)
    # Net Profit = Sale Price - Purchase Price - Marketplace Fee
    # ROI = ((Sale Price - Purchase Price - Marketplace Fee) / Purchase Price) * 100
    # Days Held = Sale Date - Purchase Date
    
    for row in range(2, 12):  # Add formulas for first 10 rows
        # Net Profit (column H)
        sheet.cell(row=row, column=8).value = f"=E{row}-C{row}-G{row}"
        sheet.cell(row=row, column=8).number_format = '$#,##0.00'
        
        # ROI (column I)
        sheet.cell(row=row, column=9).value = f"=IF(C{row}>0,((E{row}-C{row}-G{row})/C{row})*100,0)"
        sheet.cell(row=row, column=9).number_format = '0.00"%"'
        
        # Days Held (column J)
        sheet.cell(row=row, column=10).value = f"=IF(AND(D{row}<>\"\",B{row}<>\"\"),D{row}-B{row},\"\")"
        sheet.cell(row=row, column=10).number_format = '0'
    
    # Add sample data
    sample_data = [
        ["sold-example.com", "2025-01-10", 2.99, "2025-02-15", 55.00, "Flippa", 5.50],
        ["quick-brand.io", "2025-01-05", 8.99, "2025-02-20", 120.00, "Brandpa", 12.00],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if col_num not in [8, 9, 10]:  # Skip formula columns
                cell.value = value
            cell.border = border
            if col_num in [3, 5, 7]:  # Price columns
                cell.number_format = '$#,##0.00'
    
    # Freeze top row
    sheet.freeze_panes = 'A2'


def setup_summary_sheet(sheet):
    """Set up the Summary sheet."""
    # Style configuration
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")
    
    # Title
    sheet.cell(row=1, column=1).value = "Domain Flipping Portfolio Summary"
    sheet.cell(row=1, column=1).font = title_font
    
    # Current Inventory Section
    sheet.cell(row=3, column=1).value = "Current Inventory"
    sheet.cell(row=3, column=1).font = header_font
    sheet.cell(row=3, column=1).fill = header_fill
    
    metrics = [
        ("Total Domains Owned", "=COUNTA('Domain Inventory'!A:A)-1"),
        ("Total Investment", "=SUM('Domain Inventory'!C:C)"),
        ("Target Portfolio Value", "=SUM('Domain Inventory'!F:F)"),
        ("Potential Profit", "=C6-C5"),
    ]
    
    row = 4
    for label, formula in metrics:
        sheet.cell(row=row, column=1).value = label
        sheet.cell(row=row, column=2).value = formula
        if row > 4:
            sheet.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1
    
    # Sales Performance Section
    sheet.cell(row=row + 1, column=1).value = "Sales Performance"
    sheet.cell(row=row + 1, column=1).font = header_font
    sheet.cell(row=row + 1, column=1).fill = header_fill
    
    row += 2
    sales_metrics = [
        ("Total Domains Sold", "=COUNTA('Sold Domains'!A:A)-1"),
        ("Total Revenue", "=SUM('Sold Domains'!E:E)"),
        ("Total Costs", "=SUM('Sold Domains'!C:C)+SUM('Sold Domains'!G:G)"),
        ("Net Profit", "=C11-C12"),
        ("Average ROI", "=AVERAGE('Sold Domains'!I:I)"),
        ("Average Days to Sale", "=AVERAGE('Sold Domains'!J:J)"),
    ]
    
    for label, formula in sales_metrics:
        sheet.cell(row=row, column=1).value = label
        sheet.cell(row=row, column=2).value = formula
        if row in [11, 12, 13]:
            sheet.cell(row=row, column=2).number_format = '$#,##0.00'
        elif row == 14:
            sheet.cell(row=row, column=2).number_format = '0.00"%"'
        elif row == 15:
            sheet.cell(row=row, column=2).number_format = '0'
        row += 1
    
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 20
    
    # Add instructions
    row += 2
    sheet.cell(row=row, column=1).value = "Instructions:"
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    instructions = [
        "1. Add purchased domains to the 'Domain Inventory' sheet",
        "2. When a domain is sold, move it to the 'Sold Domains' sheet",
        "3. Formulas will automatically calculate profits and ROI",
        "4. This summary updates automatically based on your data",
        "5. Target prices help you set realistic sale expectations"
    ]
    
    row += 1
    for instruction in instructions:
        sheet.cell(row=row, column=1).value = instruction
        row += 1


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description='Create domain pricing spreadsheet template'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='domain_pricing_template.xlsx',
        help='Output filename (default: domain_pricing_template.xlsx)'
    )
    
    args = parser.parse_args()
    
    create_pricing_template(args.output)


if __name__ == '__main__':
    main()
